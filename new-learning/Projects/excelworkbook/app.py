import openpyxl as xl
from openpyxl import chart
from openpyxl.chart import BarChart, Reference, bar_chart


wb = xl.load_workbook("transactions.xlsx")
sheet = wb['Sheet1']
# cell = sheet['a1']
# cell = sheet.cell(1,1) access cell

# cell = sheet.cell(1, 2)  # first row second column
# cell = sheet.cell(1, 3)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)  # 2 row 3 column
    correct_price = cell.value * 0.9
    correct_price_cell = sheet.cell(row, 4)
    correct_price_cell.value = correct_price
    # correct_price_cell.value = str('$' + str(correct_price))

values = Reference(sheet, min_row=2, max_row=sheet.max_row,
                   min_col=4, max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'b8')

wb.save('updatedFile.xlsx')
